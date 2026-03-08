"""
Excel Import Script
Imports existing job tracking data from Martin's Excel spreadsheet
into the JobPilot database.

Usage:
    python scripts/import_excel.py --file path/to/tracking.xlsx
"""

import argparse
import sys
from pathlib import Path
from datetime import datetime

# Add parent to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl
from backend.db.database import get_sync_connection, DB_PATH, init_database
from backend.agents.discovery import generate_fingerprint


def import_excel(file_path: str):
    """Import jobs from Excel tracking sheet."""
    path = Path(file_path)
    if not path.exists():
        print(f"Error: File not found: {path}")
        sys.exit(1)

    # Initialize DB if needed
    from backend.db.database import DB_PATH, init_database
    if not DB_PATH.exists():
        init_database()

    wb = openpyxl.load_workbook(str(path))
    ws = wb.active

    # Get headers from first row
    headers = [cell.value for cell in ws[1] if cell.value]
    print(f"Found headers: {headers}")

    # Map expected columns (flexible matching)
    col_map = {}
    for i, h in enumerate(headers):
        h_lower = h.lower().strip() if h else ""
        if "title" in h_lower or "role" in h_lower or "job" in h_lower:
            col_map["title"] = i
        elif "company" in h_lower:
            col_map["company"] = i
        elif "link" in h_lower or "url" in h_lower:
            col_map["url"] = i
        elif "status" in h_lower:
            col_map["status"] = i
        elif "date" in h_lower or "applied" in h_lower:
            col_map["date"] = i
        elif "note" in h_lower:
            col_map["notes"] = i

    print(f"Column mapping: {col_map}")

    if "title" not in col_map or "company" not in col_map:
        print("Error: Could not find 'title' and 'company' columns")
        sys.exit(1)

    conn = get_sync_connection()
    imported = 0
    skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[col_map.get("title", 0)]:
            continue

        title = str(row[col_map["title"]]).strip()
        company = str(row[col_map["company"]]).strip().lstrip('\n').strip()
        url = str(row[col_map.get("url", 0)] or "").strip() if "url" in col_map else ""
        status_raw = str(row[col_map.get("status", 0)] or "").strip().lower() if "status" in col_map else "applied"
        date_val = row[col_map.get("date", 0)] if "date" in col_map else None
        notes_val = str(row[col_map.get("notes", 0)] or "").strip() if "notes" in col_map else ""

        # Map status values
        status_map = {
            "applied": "applied",
            "screening": "screening",
            "screener": "screening",
            "phone screen": "phone_screen",
            "interview": "interview_1",
            "interviews": "interview_1",
            "selected": "offer",
            "not selected": "rejected",
            "rejected": "rejected",
            "withdrawn": "withdrawn",
            "unknown": "applied",  # Default unknown to applied
            "none": "applied",
        }
        status = status_map.get(status_raw, "applied")

        # Parse date
        applied_date = None
        if date_val:
            if isinstance(date_val, datetime):
                applied_date = date_val.strftime("%Y-%m-%d")
            else:
                try:
                    applied_date = str(date_val)[:10]
                except Exception:
                    pass

        fingerprint = generate_fingerprint(title, company)

        # Check for duplicate
        existing = conn.execute(
            "SELECT id FROM jobs WHERE fingerprint = ?", (fingerprint,)
        ).fetchone()

        if existing:
            skipped += 1
            continue

        # Insert job
        cursor = conn.execute("""
            INSERT INTO jobs
                (fingerprint, title, company_name, source, source_url, status)
            VALUES (?, ?, ?, 'manual', ?, ?)
        """, (fingerprint, title, company, url or "imported", status))

        job_id = cursor.lastrowid

        # Create application record
        conn.execute("""
            INSERT INTO applications (job_id, status, applied_date, notes)
            VALUES (?, ?, ?, ?)
        """, (job_id, status, applied_date, notes_val or None))

        # Log activity
        conn.execute("""
            INSERT INTO activity_log (event_type, job_id, details)
            VALUES ('job_discovered', ?, 'Imported from Excel')
        """, (job_id,))

        imported += 1

    conn.commit()
    conn.close()

    print(f"\nImport complete: {imported} imported, {skipped} skipped (duplicates)")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Import Excel job tracking data")
    parser.add_argument("--file", required=True, help="Path to Excel file")
    args = parser.parse_args()
    import_excel(args.file)
